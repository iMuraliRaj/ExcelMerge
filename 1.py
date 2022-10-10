import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment

# work_book = Workbook("D:\\check\\RN\\trendnew")
# work_sheet = work_book.active
work_book = openpyxl.load_workbook("E:\\File\\Trend.xlsx")
work_sheet = work_book["Sheet1"]

numberOfRows = work_sheet.max_row

listOne = []


for i in range(2,numberOfRows+1):
    a=work_sheet.cell(i,1)
    listOne.append(a.value)


for i in range(0, len(listOne)):
    for j in range(i+1, len(listOne)):
        if(listOne[i] == listOne[j]):
            print(listOne[j]);

# cells to merge
work_sheet.merge_cells('A2:A7')

cell = work_sheet.cell(row=1, column=1)
# value of cell
cell.value = 'quick fox jumps over the lazy dog'
# aligment of data in cell
cell.alignment = Alignment(horizontal='center', vertical='center')

# save the workbook
work_book.save('E:\\File\\Trend1.xlsx')